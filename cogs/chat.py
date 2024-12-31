# cogs/chat.py
from db import db  # Import db from db.py instead of app.py
import os
import json
from flask import Blueprint, request, jsonify, session, send_from_directory  # **ADDED send_from_directory**
import openai
import copy
from .web_search import WebSearchCog
from .code_files import CodeFilesCog
from datetime import datetime
from docx import Document
from openpyxl import load_workbook
import uuid
import tiktoken
import time
from PyPDF2 import PdfReader
from werkzeug.utils import secure_filename  # **ADDED secure_filename**


# Define Database Models
class Conversation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    session_id = db.Column(db.String(100), nullable=False)
    title = db.Column(db.String(255), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    messages = db.relationship('Message', backref='conversation', lazy=True)

class Message(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    conversation_id = db.Column(db.Integer, db.ForeignKey('conversation.id'), nullable=False)
    role = db.Column(db.String(50), nullable=False)
    content = db.Column(db.Text, nullable=False)
    # Optional: Add file fields if desired
    # file_url = db.Column(db.String(500), nullable=True)      # **OPTIONAL**
    # file_name = db.Column(db.String(255), nullable=True)     # **OPTIONAL**
    # file_type = db.Column(db.String(100), nullable=True)     # **OPTIONAL**
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

class UploadedFile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    session_id = db.Column(db.String(100), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    file_url = db.Column(db.String(500), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

class ChatBlueprint:
    def __init__(self, app_instance, flask_app):
        self.bp = Blueprint("chat_blueprint", __name__)
        
        # Initialize OpenAI client properly
        openai.api_key = os.getenv('OPENAI_KEY')
        self.client = openai  # Assign the openai module as the client

        # Pass the OpenAI client to WebSearchCog
        self.web_search_cog = WebSearchCog(openai_client=self.client)
        self.code_files_cog = CodeFilesCog()
        
        self.google_key = os.getenv('GOOGLE_API_KEY')
        self.app_instance = app_instance

        # **ADDED** Ensure 'uploads' directory exists
        # self.upload_folder = os.path.join(self.app_instance.root_path, 'uploads')
        self.upload_folder = os.path.join(flask_app.instance_path, 'uploads')

        if not os.path.exists(self.upload_folder):
            os.makedirs(self.upload_folder)

        self.add_routes()

    def add_routes(self):
        @self.bp.route("/chat", methods=["POST"])
        def chat():
            # Ensure session has a unique session_id

            if not os.path.exists(self.upload_folder):
                os.makedirs(self.upload_folder)
                print(f"Uploads directory created at: {self.upload_folder}")
            else:
                print(f"Uploads directory already exists: {self.upload_folder}")


            if 'session_id' not in session:
                session['session_id'] = str(uuid.uuid4())
            session_id = session['session_id']
            
            file_content = ''
            file = None
            additional_instructions = (
                "You are an AI assistant that generates structured and easy-to-read responses.  \n"
                "Provide responses using correct markdown format. It is critical that markdown format is used with nothing additional.  \n"
                "Use headings (e.g., ## Section Title), numbered lists, and bullet points to format output.  \n"
                "Ensure sufficient line breaks between sections to improve readability. Generally, limit responses to no more than 1500 tokens."
            )
            if request.content_type.startswith('multipart/form-data'):
                # **HANDLE FILE UPLOAD**

                # Extract form data
                user_message = request.form.get("message", "")
                model = request.form.get("model", "gpt-4o")
                temperature = float(request.form.get("temperature", 0.7))
                system_prompt = request.form.get(
                    "system_prompt",
                    "You are a USMC AI agent. Provide relevant responses."
                )


                file = request.files.get("file", None)

                if file:
                    # **SECURELY SAVE THE FILE**
                    filename = secure_filename(file.filename)
                    # filename = file.filename
                    unique_filename = f"{uuid.uuid4()}_{filename}"
                    # unique_filename = filename
                    file_path = os.path.join(self.upload_folder, unique_filename)
                    file.save(file_path)

                    print(f"File saved at: {file_path}")

                    # **INSERT INTO DATABASE**
                    uploaded_file = UploadedFile(
                        session_id=session_id,
                        filename=filename,
                        file_url=f"/uploads/{unique_filename}"  # Assuming you serve files from this path
                    )
                    db.session.add(uploaded_file)
                    db.session.commit()

                    # Reset the file pointer
                    file.seek(0)
                    WORD_LIMIT = 50000

                    if file and file.content_type == 'application/pdf':
                        try:
                            reader = PdfReader(file)
                            file_content = ""
                            for page in reader.pages:
                                file_content += page.extract_text() or ""
                            
                            # Truncate to 50,000 words
                            words = file_content.split()
                            if len(words) > WORD_LIMIT:
                                file_content = ' '.join(words[:WORD_LIMIT]) + "\n\n[Text truncated after 50,000 words.]"

                            if not file_content.strip():
                                file_content = "Unable to extract text from this PDF."

                        except Exception as e:
                            print("Error reading PDF:", e)
                            file_content = "Error processing PDF file."
                    
                    # **PROCESS WORD FILES (DOCX)**
                    elif file.content_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document', 'application/msword']:
                        try:
                            doc = Document(file_path)
                            file_content = "\n".join([p.text for p in doc.paragraphs])

                            words = file_content.split()
                            if len(words) > WORD_LIMIT:
                                file_content = ' '.join(words[:WORD_LIMIT]) + "\n\n[Text truncated after 50,000 words.]"

                        except Exception as e:
                            print("Error reading DOCX:", e)
                            file_content = "Error processing Word file."

                    # **PROCESS EXCEL FILES (XLSX)**
                    elif file.content_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel']:
                        try:
                            wb = load_workbook(file_path)
                            sheet = wb.active
                            file_content = ""
                            for row in sheet.iter_rows(values_only=True):
                                file_content += ' '.join(str(cell) for cell in row if cell is not None) + "\n"

                            words = file_content.split()
                            if len(words) > WORD_LIMIT:
                                file_content = ' '.join(words[:WORD_LIMIT]) + "\n\n[Text truncated after 50,000 words.]"

                        except Exception as e:
                            print("Error reading Excel file:", e)
                            file_content = "Error processing Excel file."

                    # **DEFAULT TO TEXT FILES OR OTHER PLAIN FORMATS**
                    else:
                        file_content = file.read().decode('utf-8', errors='ignore')

                    # **GENERATE FILE URL**
                    file_url = f"/uploads/{unique_filename}"
                    file_type = file.content_type
                    
                else:
                    file_url = None
                    file_type = None
                    file_content = ''
                # print()
                # print('file_content', file_content)
            else:
                # **HANDLE JSON REQUEST**
                data = request.get_json()
                user_message = data.get("message", "")
                model = data.get("model", "gpt-4o")
                temperature = float(data.get("temperature", 0.7))
                system_prompt = data.get(
                    "system_prompt",
                    "You are a USMC AI agent. Provide relevant responses."
                )
                file_url = None
                file_type = None

            if not user_message and not file_url:
                return jsonify({"error": "No message or file provided"}), 400


            # Check if there's a current conversation
            if 'current_conversation_id' not in session:
                # Create a new conversation
                title = "New Conversation"
                new_convo = Conversation(
                    session_id=session_id,
                    title=title
                )
                db.session.add(new_convo)
                db.session.commit()
                session['current_conversation_id'] = new_convo.id

            conversation_id = session.get('current_conversation_id')
            conversation = Conversation.query.get(conversation_id)

            if not conversation or conversation.session_id != session_id:
                return jsonify({"error": "Conversation not found or unauthorized"}), 404

            # Get messages for the conversation
            messages = Message.query.filter_by(conversation_id=conversation_id).order_by(Message.timestamp).all()
            conversation_history = [{"role": msg.role, "content": msg.content} for msg in messages]

            # Add system prompt if first message
            if not conversation_history:
                conversation_history.append({
                    "role": "system",
                    "content": system_prompt + '\n' + additional_instructions
                })

            # Append user message
            if file_url:
                user_message_entry = {
                    "role": "user",
                    "content": user_message,
                    "fileUrl": file_url,
                    "fileName": filename,
                    "fileType": file_type
                }
            else:
                user_message_entry = {"role": "user", "content": user_message}

            conversation_history.append(user_message_entry)

            print('conversation_history', conversation_history)

            # Save the user message
            self.save_messages(conversation_id, "user", user_message)

            # **ADDED** Save file info in the message (Optional)
            if file_url:
                # Assuming you want to store file information in the message content as JSON or formatted text
                # Here, we'll append the file URL to the message content
                last_message = Message.query.filter_by(conversation_id=conversation_id, role="user").order_by(Message.timestamp.desc()).first()
                if last_message:
                    last_message.content = f"{user_message}\n\n[Uploaded File]({file_url})"
                    db.session.commit()

            # Analyze user intent
            if file:
                print('@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@')
                print()
                print('If file')
                print()
                print('@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@')
                intent = self.analyze_user_intent(user_message + "A file was uploaded. The text is likely refering to this file.", conversation_history, session_id)
            else:
                intent = self.analyze_user_intent(user_message, conversation_history, session_id)

            print()
            print('intent', intent)
            print()

            # Handle different intents
            assistant_reply = ""
            try:
                if intent.get("image_generation", False):
                    # Generate Image using DALL-E 3
                    prompt = intent.get("image_prompt", "")
                    if prompt:
                        image_url = self.generate_image(prompt)
                        assistant_reply = f"![Generated Image]({image_url})"
                        conversation_history.append({"role": "assistant", "content": assistant_reply})
                        # Save messages
                        self.save_messages(conversation_id, "assistant", assistant_reply)
                    else:
                        assistant_reply = "No image prompt provided."
                        conversation_history.append({"role": "assistant", "content": assistant_reply})
                        self.save_messages(conversation_id, "assistant", assistant_reply)
                    return jsonify({
                        "user_message": user_message,
                        "assistant_reply": assistant_reply,
                        "conversation_history": conversation_history,
                        "intent": intent,
                        "fileUrl": file_url,         # **ADDED**
                        "fileName": filename if file_url else None,   # **ADDED**
                        "fileType": file_type if file_url else None    # **ADDED**
                    })
                                
                elif intent.get("file_intent", False):
                    file_id = intent.get("file_id")

                    if not file_id:
                        assistant_reply = "No file ID provided."
                    else:
                        # Retrieve the file from the database
                        uploaded_file = UploadedFile.query.get(file_id)
                        
                        if uploaded_file:
                            file_path = os.path.join(self.upload_folder, uploaded_file.filename)
                            
                            # Check if the file exists
                            if os.path.exists(file_path):
                                try:
                                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                                        file_content = f.read()
                                    
                                    # Truncate content if necessary
                                    max_tokens = 50000
                                    words = file_content.split()
                                    if len(words) > max_tokens:
                                        file_content = ' '.join(words[:max_tokens]) + "\n\n[Content truncated...]"
                                    
                                    assistant_reply = f"The content of the file '{uploaded_file.filename}' is:\n\n{file_content}"
                                
                                except Exception as e:
                                    print(f"Error reading file: {e}")
                                    assistant_reply = f"Could not read the file '{uploaded_file.filename}'."
                            else:
                                print(f"File not found: {file_path}")
                                assistant_reply = f"File '{uploaded_file.filename}' not found. It may have been deleted or the session expired."
                                
                                # Optional: Remove stale DB entry
                                db.session.delete(uploaded_file)
                                db.session.commit()
                                print(f"Removed stale entry for '{uploaded_file.filename}' from the database.")
                        
                        else:
                            assistant_reply = f"File with ID {file_id} not found in the database."



                    # Add the assistant reply to the conversation history
                    conversation_history.append({"role": "assistant", "content": assistant_reply})
                    self.save_messages(conversation_id, "assistant", assistant_reply)

                elif intent.get("code_intent", False):
                    # Handle code-related intents
                    code_content = self.code_files_cog.get_all_code_files_content()
                    if code_content:
                        # Append code content to the system prompt
                        temp_conversation = copy.deepcopy(conversation_history)
                        temp_conversation[0]['content'] += "\n\n" + code_content
                        temp_conversation[-1]['content'] += '\n\nYou have been supplemented with information from your code base to answer this query.'
                        intent['code_intent'] = True  # Ensure intent shows True
                    else:
                        assistant_reply = "No code files found to provide."
                        conversation_history.append({"role": "assistant", "content": assistant_reply})
                        self.save_messages(conversation_id, "assistant", assistant_reply)
                        return jsonify({
                            "user_message": user_message,
                            "assistant_reply": assistant_reply,
                            "conversation_history": conversation_history,
                            "intent": intent,
                            "fileUrl": file_url,         # **ADDED**
                            "fileName": filename if file_url else None,   # **ADDED**
                            "fileType": file_type if file_url else None    # **ADDED**
                        })

                elif intent.get("internet_search", False):
                    # Handle internet search
                    query = user_message  # Or extract a specific part of the message
                    search_content = self.web_search_cog.web_search(query, conversation_history)
                    sys_search_content = (
                        '\nDo not say "I am unable to browse the internet," because you have information directly retrieved from the internet. '
                        'Give a confident answer based on this. Only use the most relevant and accurate information that matches the User Query. '
                        'Always include the source with the provided url as [source](url)'
                    )

                    # Integrate the internet content into the system prompt
                    temp_conversation = copy.deepcopy(conversation_history)
                    temp_conversation[0]['content'] += sys_search_content
                    temp_conversation[-1]['content'] = (
                        '\n\nYou are being supplemented with the following information from the internet to answer user query. '
                        f"Internet Content:\n***{search_content}***\n\nUser Query:\n***{user_message}***"
                    )
                else:
                    print('in else: conversation_history', conversation_history)
                    temp_conversation = copy.deepcopy(conversation_history)
                    # temp_conversation[-1]['content'] = (f'\n{file_content}\n') + temp_conversation[-1]['content']



                def trim_conversation(temp_conversation, max_tokens=50000):
                    # print('Before trim: temp_conversation', temp_conversation)
                    
                    encoding = tiktoken.encoding_for_model("gpt-4o")
                    total_tokens = 0
                    trimmed = []
                    
                    for message in reversed(temp_conversation):
                        message_tokens = len(encoding.encode(json.dumps(message)))
                        if total_tokens + message_tokens > max_tokens:
                            break
                        trimmed.insert(0, message)
                        total_tokens += message_tokens
                    
                    # Ensure at least one message is included
                    if not trimmed and temp_conversation:
                        trimmed.append(temp_conversation[-1])
                    
                    # print('After trim: temp_conversation', trimmed)
                    return trimmed

                # Trim the conversation if it exceeds 8000 tokens
                
                temp_conversation = trim_conversation(temp_conversation, 50000)

                # Regular Chat Response
                print('***************************************************************************************************************')
                print()
                print('temp_conversation', temp_conversation)
                print('***************************************************************************************************************')
                print()
                start_time = time.time()
                assistant_reply = self.generate_chat_response(temp_conversation, model, temperature)
                end_time = time.time()
                print()
                print("generate_chat_response took %s seconds", end_time - start_time)
                print()
                print()
                print('Final Response:', assistant_reply)
                print()
                print()
                conversation_history.append({"role": "assistant", "content": assistant_reply})
                # Save messages

                self.save_messages(conversation_id, "assistant", assistant_reply)

                return jsonify({
                    "user_message": user_message,
                    "assistant_reply": assistant_reply,
                    "conversation_history": conversation_history,
                    "intent": intent,
                    "fileUrl": file_url,         # **ADDED**
                    "fileName": filename if file_url else None,   # **ADDED**
                    "fileType": file_type if file_url else None    # **ADDED**
                })

            except Exception as e:
                return jsonify({"error": str(e)}), 500

        @self.bp.route("/uploads/<path:filename>", methods=["GET"])
        def uploaded_file(filename):
            # Ensure the request is part of the current session
            session_id = session.get('session_id', None)
            if not session_id:
                return jsonify({"error": "Unauthorized access"}), 403
            
            # Verify that the requested file belongs to the current session
            file_entry = UploadedFile.query.filter_by(session_id=session_id, file_url=f"/uploads/{filename}").first()
            if not file_entry:
                return jsonify({"error": "File not found"}), 404
            
            # Serve the file
            return send_from_directory(self.upload_folder, filename)

            
        @self.bp.route("/conversations", methods=["GET"])
        def get_conversations():
            # Fetch recent conversations for the current session
            session_id = session.get('session_id', 'unknown_session')
            conversations = Conversation.query.filter_by(session_id=session_id).order_by(Conversation.timestamp.desc()).limit(10).all()
            convo_list = [{
                "id": convo.id,
                "title": convo.title,
                "timestamp": convo.timestamp.isoformat()
            } for convo in conversations]
            return jsonify({"conversations": convo_list})

        @self.bp.route("/conversations/<int:conversation_id>", methods=["GET"])
        def get_conversation(conversation_id):
            # Fetch a specific conversation's messages
            conversation = Conversation.query.get(conversation_id)
            if not conversation:
                return jsonify({"error": "Conversation not found"}), 404
            session_id = session.get('session_id', 'unknown_session')
            if conversation.session_id != session_id:
                return jsonify({"error": "Unauthorized access"}), 403
            messages = Message.query.filter_by(conversation_id=conversation_id).order_by(Message.timestamp).all()
            conversation_history = [{"role": msg.role, "content": msg.content} for msg in messages]
            return jsonify({"conversation_history": conversation_history})

        @self.bp.route("/conversations/new", methods=["POST"])
        def new_conversation():
            session_id = session.get('session_id', 'unknown_session')
            data = request.get_json()
            title = data.get('title', 'New Conversation')
            new_convo = Conversation(
                session_id=session_id,
                title=title
            )
            db.session.add(new_convo)
            db.session.commit()
            session['current_conversation_id'] = new_convo.id
            return jsonify({"conversation_id": new_convo.id, "title": new_convo.title, "timestamp": new_convo.timestamp.isoformat()})

    def save_messages(self, conversation_id, role, content):
        """Save a message to the database."""
        msg = Message(
            conversation_id=conversation_id,
            role=role,
            content=content
        )
        db.session.add(msg)
        db.session.commit()

    def analyze_user_intent(self, user_input, conversation_hist, session_id=None):
        """Analyze user intent using OpenAI and return a JSON object."""

        # Fetch the list of uploaded files for the current session
        uploaded_files = UploadedFile.query.filter_by(session_id=session_id).all()

        file_list = "\n".join([f"File ID: {file.id}, Filename: {file.filename}" for file in uploaded_files])

        
        analysis_prompt = [
            {
                'role': 'system', 
                'content': (
                    'As an AI assistant, analyze the user input, including the last 5 user queries, and output a JSON object with the following keys:\n'
                    '- "image_generation": (boolean)\n'
                    '- "image_prompt": (string)\n'
                    '- "internet_search": (boolean)\n'
                    '- "file_intent": (boolean)\n'
                    '- "file_id": (string)\n'
                    '- "active_users": (boolean)\n'
                    '- "code_intent": (boolean)\n'
                    '- "rand_num": (list)\n\n'
                    'Respond with only the JSON object and no additional text.\n\n'
                    'Guidelines:\n'
                    '1. **image_generation** should be True only when an image is requested. Example: "Can you show me a USMC officer saluting?"\n'
                    '2. **image_prompt** should contain the prompt for image generation if **image_generation** is True.\n'
                    '3. **internet_search** should be True when the user asks for information that might require an internet search. If asking about an uploaded file, set to False.\n'
                    f'4. **file_intent** should be True when the user asks for information about a file that has been uploaded. Set to True is asked about one of these files:\n{file_list}\n'
                    '5. **file_id** should contain the file_id for the requested file if **file_intent** is True. Detect file references in the format "FILE:<id>".\n'
                    '6. **active_users** should be True if there is a question about the most active users.\n'
                    '7. **code_intent** should be True when the user is asking about code-related queries or commands starting with "!".\n'
                    '8. **rand_num** should contain [lowest_num, highest_num] if the user requests a random number within a range.\n\n'
                    'Respond in JSON format.\nIMPORTANT: Boolean values only: True or False.'
                )
            },
            {'role': 'user', 'content': f"User input: '{user_input}'\n\nDetermine the user's intent and required actions."}
        ]
        # Include the last 5 messages for context
        analysis_prompt.extend(conversation_hist[-5:])

        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",  # Correct model name
                messages=analysis_prompt,
                max_tokens=300,
                temperature=0
            )

            print("Full Response:", response)

            intent_json = response.choices[0].message.content.strip()

            # Handle markdown-wrapped JSON
            if intent_json.startswith("```json"):
                intent_json = intent_json[7:-3].strip()  # Strip off ```json and ```
            # Ensure the response is valid JSON
            intent = json.loads(intent_json)

            # Additional logic to parse file_id if file_intent is detected
            if intent.get("file_intent", False):
                # Extract file_id from user_input, assuming format "FILE:<id>"
                import re
                match = re.search(r"FILE:(\d+)", user_input)
                if match:
                    intent["file_id"] = match.group(1)

            return intent
        except Exception as e:
            print()
            print(f'Error in analyzing user intent: {e}')
            print()
            # Return default intent if analysis fails
            return {
                "image_generation": False,
                "image_prompt": "",
                "internet_search": False,
                "file_intent": False,
                "file_id": "",
                "active_users": False,
                "code_intent": False,
                "rand_num": []
            }

    def generate_image(self, prompt):
        """Generate an image using OpenAI's DALL-E 3 and return the image URL."""
        try:
            image_response = self.client.images.generate(
                model="dall-e-3",
                prompt=prompt,
                size="1024x1024",
                n=1
            )
            image_url = image_response.data[0].url
            return image_url
        except Exception as e:
            print()
            print(f'Error in image generation: {e}')
            print()
            return "Error generating image."

    def generate_chat_response(self, conversation, model, temperature):
        """Generate a chat response using OpenAI's ChatCompletion."""
        try:
            response = self.client.chat.completions.create(
                model=model,
                messages=conversation,
                max_tokens=2000,  # You might want to set this based on requirements
                temperature=temperature
            )
            assistant_reply = response.choices[0].message.content
            return assistant_reply
        except Exception as e:
            print()
            print(f'Error in chat response generation: {e}')
            print()
            return "Error generating response."
